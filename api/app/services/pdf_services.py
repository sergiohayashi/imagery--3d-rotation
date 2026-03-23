import os
import tempfile

from pypdf import PdfReader
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

TMP_UPLOAD_FOLDER = tempfile.gettempdir()


class FileService:

    @staticmethod
    async def extract_content(file):
        filename = secure_filename(file.filename)
        file_path = os.path.join(TMP_UPLOAD_FOLDER, filename)

        # Save the uploaded file to a temporary file
        with open(file_path, "wb") as out_file:
            while content := await file.read(1024):  # Read chunks of 1024 bytes
                out_file.write(content)

        # Extract content based on file type
        if filename.endswith(".pdf"):
            return FileService.extract_text_from_pdf(file_path)
        elif filename.endswith(".xlsx"):
            return FileService.extract_data_from_xlsx(file_path)
        else:
            return None

    @staticmethod
    def extract_text_from_pdf(file_path):
        text = ""
        reader = PdfReader(file_path)
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text

    @staticmethod
    def extract_data_from_xlsx(file_path):
        workbook = load_workbook(file_path)

        text = ""
        for sheet in workbook.sheetnames:
            text += f"** Sheet: {sheet} **\n"
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                row_text = "\t".join(
                    [str(cell) if cell is not None else "" for cell in row]
                )
                text += row_text + "\n"
            text += "\n"  # Add an extra newline between sheets
        return text

        # sheet = workbook.active
        # text = ""
        # for row in sheet.iter_rows(values_only=True):
        #     row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
        #     text += row_text + "\n"
        # return text
